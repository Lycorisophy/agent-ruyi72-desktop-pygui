#!/usr/bin/env python3
"""
Excel Spreadsheet Writer for OpenClaw
Creates professional .xlsx spreadsheets
"""

import argparse
import json
import sys
from pathlib import Path
from datetime import datetime


def create_excel_document(content_dict: dict, output_path: str) -> dict:
    """
    创建 Excel 表格
    
    Args:
        content_dict: 表格内容配置
        output_path: 输出文件路径
    
    Returns:
        创建结果
    """
    results = {
        'success': False,
        'path': output_path,
        'message': ''
    }
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        results['message'] = """
❌ 缺少依赖库 openpyxl

请安装依赖：
pip install openpyxl

或者先使用 md-writer 技能创建 Markdown 表格。
"""
        return results
    
    wb = Workbook()
    
    # 默认工作表
    ws = wb.active
    ws.title = "Sheet1"
    
    # 设置标题行样式
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    
    # 创建工作表
    sheet_data = content_dict.get('sheets', [content_dict])
    
    for idx, sheet_info in enumerate(sheet_data):
        if idx == 0:
            ws = wb.active
        else:
            ws = wb.create_sheet(title=sheet_info.get('name', f'Sheet{idx+1}'))
        
        # 标题
        if 'title' in sheet_info:
            ws['A1'] = sheet_info['title']
            ws.merge_cells('A1:' + get_column_letter(len(sheet_info.get('headers', [['A']]))) + '1')
            ws['A1'].font = Font(bold=True, size=16)
            ws['A1'].alignment = Alignment(horizontal='center')
        
        # 表头
        headers = sheet_info.get('headers', [])
        if headers:
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col_idx, value=header)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
        
        # 数据行
        rows = sheet_info.get('rows', [])
        for row_idx, row_data in enumerate(rows, 3):
            for col_idx, cell_data in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_data)
                cell.alignment = Alignment(vertical='center')
        
        # 调整列宽
        if headers:
            for col_idx, header in enumerate(headers, 1):
                column_letter = get_column_letter(col_idx)
                max_length = max(len(str(header)), max((len(str(row.get(col_idx-1, ''))) for row in rows) or 0))
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # 添加筛选
        if headers:
            ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{len(rows) + 2}"
    
    # 多工作表支持
    for sheet_name in content_dict.get('worksheets', []):
        ws_new = wb.create_sheet(title=sheet_name)
        ws_new['A1'] = sheet_name
    
    # 保存文件
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    
    results['success'] = True
    results['message'] = f"""
✅ Excel 表格创建成功！

📄 文件路径: {output_path}
📏 文件大小: {output_path.stat().st_size / 1024:.1f} KB

可直接在 Microsoft Excel 或其他兼容软件中打开。
"""
    return results


def format_results(results: dict) -> str:
    """格式化结果"""
    return results.get('message', '❌ 创建失败')


def main():
    parser = argparse.ArgumentParser(
        description='Excel Spreadsheet Writer for OpenClaw',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python excel_writer.py --title "数据表" --content '{"headers": ["A", "B"], "rows": [["1", "2"]]}'
  python excel_writer.py --output "data.xlsx" --data '{"sheets": [{"name": "统计", "headers": ["项目", "数值"], "rows": [["A", 100]]}]}'
        """
    )
    
    parser.add_argument('--output', '-o', default='spreadsheet.xlsx', help='Output file path')
    parser.add_argument('--title', '-t', default='新表格', help='Spreadsheet title')
    parser.add_argument('--content', '-c', default='{}', help='Content as JSON')
    parser.add_argument('--data', '-d', help='Full spreadsheet data as JSON')
    
    args = parser.parse_args()
    
    if args.data:
        content_dict = json.loads(args.data)
    else:
        content_dict = json.loads(args.content)
    
    results = create_excel_document(content_dict, args.output)
    print(format_results(results))


if __name__ == '__main__':
    main()
